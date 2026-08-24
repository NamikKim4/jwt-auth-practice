"""리스트 데이터를 .xlsx 파일(엑셀)로 변환하는 공용 헬퍼."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def build_excel(headers: list[str], rows: list[list]) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    ws.append(headers)
    header_fill = PatternFill(start_color="FFDCE6FF", end_color="FFDCE6FF", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    # 컬럼 너비를 대충 내용 길이에 맞춰 자동 조정
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            value = row[col_idx - 1]
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
